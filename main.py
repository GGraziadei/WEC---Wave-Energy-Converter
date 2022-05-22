from fileinput import filename
import re
from time import time
from openpyxl import load_workbook


#valoreAltezza : [ (ValoreTempo , valoreCella )]

power_matrix = dict()

filename = "Totale Futuna.xlsx"
workbook = load_workbook(filename=filename)
sheets = workbook.sheetnames
assert sheets.__contains__('CF_WEC')
index = sheets.index('CF_WEC')
sheet = workbook[sheets[index]]
assert sheet.title.__eq__('CF_WEC')

power_matrix_filename = "PowerMatrixWEC.xlsx"
power_matrix_workbook = load_workbook(filename=power_matrix_filename)
power_matrix_sheet = power_matrix_workbook.active

time_list = []

# Lettura degli intervalli temporali
for tuple in power_matrix_sheet["C2:AI2"]:
    for cell in tuple:
        time_list.append(cell.value)

# Lettura della matrice ( la matrice è salvata in un dizionario come sopra riportato )
for row in power_matrix_sheet.iter_rows(min_col=2 , max_col=power_matrix_sheet.max_column, min_row=3 , values_only=True):
    heigh_index = row[0]
    value_list = []
    index = 0
    for value in row[1:]:
        value_tuple = ( time_list[index] , value)
        value_list.append(value_tuple)
        index = index + 1
    power_matrix[heigh_index] = value_list

def get_height_key (height):
    valore_successivo = 0
    valore_precedente = 0
    for k in  power_matrix.keys(): 
        if float(k) > height:
            valore_successivo = float(k)
            break
    for k in  reversed ( power_matrix.keys() ): 
        if float(k) < height:
            valore_precedente = float(k)
            break 
    distanza_successivo = valore_successivo - height
    distanza_precedente = height - valore_precedente
    if(distanza_successivo < distanza_precedente ):
        return valore_successivo
    return valore_precedente

def get_value (time , height_value_list):
    valore_successivo = 0
    valore_precedente = 0
    for t in height_value_list: 
        if t[0] > time:
            valore_successivo = t
            break
    for t in reversed( height_value_list ): 
        if t[0] < time:
            valore_precedente = t
            break
    distanza_successivo = valore_successivo[0] - time
    distanza_precedente = time - valore_precedente[0]
    if(distanza_successivo < distanza_precedente ):
        return valore_successivo[1]
    return valore_precedente[1]

def get_production( period , height ):
    key = get_height_key(height=height)
    height_value_list = power_matrix[key]
    return get_value(time = period , height_value_list = height_value_list)

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row ):
    period = row[1].value
    height = row[2].value
    production = get_production(period = period , height = height)
    production_cell = row[5]
    production_cell.value = production

workbook.save(filename=filename)
    
